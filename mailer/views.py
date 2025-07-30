import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from django.core.mail import send_mail

# from django.http import JsonResponse
from django.shortcuts import render
from django.db import transaction

# from django.http import HttpResponse
# from django.template import loader
from django.core.mail import send_mail, send_mass_mail
from django.template.loader import render_to_string
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from drf_spectacular.utils import (
    extend_schema,
    OpenApiParameter,
    OpenApiResponse,
    OpenApiExample,
    OpenApiRequest,
)
from drf_spectacular.types import OpenApiTypes
from .models import Email, Campaign, EmailTemplate
from .serializers import EmailSerializer, CampaignSerializer, EmailTemplateSerializer
from django.conf import settings


class CampaignCreateView(APIView):
    @extend_schema(
        request=CampaignSerializer,
        responses={201: CampaignSerializer, 400: OpenApiResponse(description="Bad Request")},
        description="Create a new campaign",
    )
    def post(self, request):
        try:
            serializer = CampaignSerializer(data=request.data)
            if serializer.is_valid():
                campaign = serializer.save()
                return Response(serializer.data, status=201)
            return Response(serializer.errors, status=400)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class UpdateCampaignView(APIView):
    def post(self, request):
        try:
            campaign_id = request.GET.get('campaign_id')
            campaign_name = request.GET.get('campaign_name')

            if not campaign_id or not campaign_name:
                return Response({"error": "campaign_id and campaign_name are required"}, status=400)

            campaign = Campaign.objects.get(id=campaign_id)
            campaign.name = campaign_name
            campaign.save()

            serializer = CampaignSerializer(campaign)
            return Response(serializer.data, status=200)
        except Campaign.DoesNotExist:
            return Response({"error": "Campaign not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class DeleteCampaignView(APIView):
    def post(self, request):
        try:
            campaign_id = request.GET.get('campaign_id')

            if not campaign_id:
                return Response({"error": "campaign_id is required"}, status=400)

            campaign = Campaign.objects.get(id=campaign_id)
            campaign.delete()

            return Response({"message": "Campaign deleted successfully"}, status=200)
        except Campaign.DoesNotExist:
            return Response({"error": "Campaign not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class ListEmailView(APIView):
    def get(self, request):
        try:
            campaign_id = request.GET.get('campaign_id')

            if not campaign_id:
                return Response({"error": "campaign_id is required"}, status=400)

            emails = Email.objects.filter(campaign_id=campaign_id)
            serializer = EmailSerializer(emails, many=True)
            return Response(serializer.data, status=200)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class XLSReaderView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        try:
            campaign_id = request.GET.get('campaign_id')
            file = request.FILES.get('file')

            if not campaign_id or not file:
                return Response({"error": "campaign_id and file are required"}, status=400)

            campaign = Campaign.objects.get(id=campaign_id)

            # Process Excel file
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            emails_created = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Assuming email is in first column
                    email_address = str(row[0]).strip()
                    name = str(row[1]).strip() if len(row) > 1 and row[1] else ""

                    Email.objects.get_or_create(
                        email_address=email_address,
                        campaign=campaign,
                        defaults={'name': name}
                    )
                    emails_created += 1

            return Response({"message": f"{emails_created} emails processed"}, status=200)
        except Campaign.DoesNotExist:
            return Response({"error": "Campaign not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class SendEmailsView(APIView):
    def post(self, request):
        try:
            campaign_id = request.GET.get('campaign_id')
            email_template = request.GET.get('email_template', 'default')

            if not campaign_id:
                return Response({"error": "campaign_id is required"}, status=400)

            campaign = Campaign.objects.get(id=campaign_id)
            emails = Email.objects.filter(campaign=campaign)

            if not emails.exists():
                return Response({"error": "No emails found for this campaign"}, status=400)

            # Send emails logic here
            sent_count = 0
            for email in emails:
                try:
                    send_mail(
                        subject=f"Campaign: {campaign.name}",
                        message=request.data.get('message', 'Default message'),
                        from_email=settings.EMAIL_HOST_USER,
                        recipient_list=[email.email_address],
                        fail_silently=False,
                    )
                    sent_count += 1
                except Exception as e:
                    continue

            return Response({"message": f"Sent {sent_count} emails"}, status=200)
        except Campaign.DoesNotExist:
            return Response({"error": "Campaign not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class DeleteEmailView(APIView):
    def post(self, request):
        try:
            email_address = request.GET.get('email_add')
            campaign_id = request.GET.get('campaign_id')

            if not email_address or not campaign_id:
                return Response({"error": "email_add and campaign_id are required"}, status=400)

            email = Email.objects.get(email_address=email_address, campaign_id=campaign_id)
            email.delete()

            return Response({"message": "Email deleted successfully"}, status=200)
        except Email.DoesNotExist:
            return Response({"error": "Email not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class UpdateEmailView(APIView):
    def post(self, request):
        try:
            email_address = request.GET.get('email_add')
            campaign_id = request.GET.get('campaign_id')
            new_name = request.data.get('name', '')

            if not email_address or not campaign_id:
                return Response({"error": "email_add and campaign_id are required"}, status=400)

            email = Email.objects.get(email_address=email_address, campaign_id=campaign_id)
            email.name = new_name
            email.save()

            serializer = EmailSerializer(email)
            return Response(serializer.data, status=200)
        except Email.DoesNotExist:
            return Response({"error": "Email not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

# 5) view for uploading emails from XLS file - user will provide campaign id and XLS file and model will be created
# 6) view for listing emails - will list all emails in the database
# 7) view for sending emails - user will provide campaign id and model will send emails to all emails in the database with that campaign id
#


class CampaignListView(APIView):
    @extend_schema(
        responses={
            200: CampaignSerializer(many=True),
            500: OpenApiResponse(
                description="Server error",
                examples={"application/json": {"error": "Error message"}},
            ),
        },
        description="List all existing campaigns.",
        examples=[
            OpenApiExample(
                "Example Response",
                value=[
                    {
                        "campaign_id": 1,
                        "campaign_name": "Holiday Sales Campaign",
                        "created_at": "2025-01-01T12:00:00Z",
                        "updated_at": "2025-01-02T08:00:00Z",
                    },
                    {
                        "campaign_id": 2,
                        "campaign_name": "Summer Discounts Campaign", 
                        "created_at": "2025-01-10T14:00:00Z",
                        "updated_at": "2025-01-12T10:30:00Z",
                    },
                ],
            )
        ],
    )
    def get(self, request):
        try:
            campaigns = Campaign.objects.all().order_by("-created_at")
            if not campaigns.exists():
                return Response([], status=200)
            serializer = CampaignSerializer(campaigns, many=True)
            return Response(serializer.data, status=200)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class CampaignCreateView(APIView):
    @extend_schema(
        request=CampaignSerializer,
        responses={
            201: CampaignSerializer,
            400: OpenApiResponse(
                description="Validation error details",
                examples={"application/json": {"error": "Validation error details"}},
            ),
            500: OpenApiResponse(
                description="Server err",
                examples={"application/json": {"error": "Error message"}},
            ),
        },
        description="Create a new campaign by providing the required details.",
        examples=[
            OpenApiExample(
                "Example Request",
                value={"campaign_name": "Black Friday Campaign"},
                request_only=True,
            ),
            OpenApiExample(
                "Example Response",
                value={
                    "campaign_id": 1,
                    "campaign_name": "AutoSAD Marketing Campaign",
                    "created_at": "2025-01-27T10:00:00Z",
                    "updated_at": "2025-01-27T10:00:00Z",
                },
                response_only=True,
            ),
        ],
    )
    def post(self, request):
        try:
            serializer = CampaignSerializer(data=request.data)
            if serializer.is_valid():
                # Check if campaign with the same name already exists
                if Campaign.objects.filter(
                    campaign_name=serializer.validated_data["campaign_name"]
                ).exists():
                    return Response(
                        {"error": "A campaign with this name already exists."},
                        status=400,
                    )

                serializer.save()
                return Response(serializer.data, status=201)

            return Response(serializer.errors, status=400)

        except Exception as e:
            return Response({"error": str(e)}, status=500)


# class CampaignUpdateView(APIView):
#     @extend_schema(
#         parameters=[
#             OpenApiParameter(
#                 name="campaign_id",
#                 type=OpenApiTypes.INT,
#                 location=OpenApiParameter.PATH,
#                 description="Campaign ID",
#             ),
#         ],
#         request=CampaignSerializer,
#         responses={
#             200: CampaignSerializer,
#             500: {"error": "Error message"}
#         },
#         description="Update an existing campaign.",
#     )
#     def put(self, request):
#         try:
#             campaign_id = request.data.get('campaign_id')
#             campaign = Campaign.objects.get(id=campaign_id)
#             serializer = CampaignSerializer(campaign, data=request.data)
#             if serializer.is_valid():
#                 serializer.save()
#                 return Response(serializer.data)
#             return Response(serializer.errors, status=400)
#         except Campaign.DoesNotExist:
#             return Response({"error": "Campaign not found"}, status=404)
#         except Exception as e:
#             return Response({"error": str(e)}, status=500)


# class CampaignDeleteView(APIView):
#     @extend_schema(
#         parameters=[
#             OpenApiParameter(
#                 name="campaign_id",
#                 type=OpenApiTypes.INT,
#                 location=OpenApiParameter.PATH,
#                 description="Campaign ID",
#             ),
#         ],
#         responses={
#             204: None,
#             500: {"error": "Error message"}
#         },
#         description="Delete an existing campaign.",
#     )
#     def delete(self, request):
#         try:
#             campaign_id = request.data.get('campaign_id')
#             campaign = Campaign.objects.get(id=campaign_id)
#             campaign.delete()
#             return Response(status=204)
#         except Campaign.DoesNotExist:
#             return Response({"error": "Campaign not found"}, status=404)
#         except Exception as e:
#             return Response({"error": str(e)}, status=500)


class XLSReaderView(APIView):
    parser_classes = [MultiPartParser]

    @extend_schema(
        request={
            "multipart/form-data": {
                "type": "object",
                "properties": {
                    "file": {
                        "type": "string",
                        "format": "binary",
                        "description": "Excel file containing email data.",
                    }
                },
            }
        },
        parameters=[
            OpenApiParameter(
                "campaign_id",
                type=int,
                location="query",
                required=True,
                description="ID of the campaign to associate the emails with.",
            )
        ],
        responses={
            201: OpenApiResponse(
                description="Emails saved successfully!",
                examples={
                    "application/json": {"message": "Emails saved successfully!"}
                },
            ),
            400: OpenApiResponse(
                description="Error in file format, missing data, or invalid campaign ID.",
                examples={"application/json": {"error": "Error message"}},
            ),
            400: OpenApiResponse(
                description="Email already exists in database.",
                examples={"application/json": {"error": "Error message"}},
            ),
            500: OpenApiResponse(
                description="Server-side error.",
                examples={"application/json": {"error": "Error message"}},
            ),
        },
        description=(
            "Upload an Excel file to save email data into the database. "
            "The file should contain 'name' and 'email_address' columns. "
            "Provide a valid campaign ID as a query parameter to associate the emails with a campaign."
        ),
        examples=[
            OpenApiExample(
                "Example Request",
                value={
                    "file": "(binary file data here)",
                },
                request_only=True,
            ),
            OpenApiExample(
                "Example Response",
                value={"message": "Emails saved successfully!"},
                response_only=True,
            ),
        ],
    )
    def post(self, request):
        file = request.FILES.get("file")
        campaign_id = request.query_params.get("campaign_id")

        # Step 1: Validate campaign ID
        if not campaign_id:
            return Response(
                {"error": "Campaign ID is required as a query parameter."}, status=400
            )

        try:
            campaign_id = int(campaign_id)
            campaign = Campaign.objects.get(campaign_id=campaign_id)
        except ValueError:
            return Response({"error": "Campaign ID must be an integer."}, status=400)
        except Campaign.DoesNotExist:
            return Response(
                {"error": "Invalid Campaign ID. Campaign not found."}, status=400
            )

        # Step 2: Validate file upload
        if not file:
            return Response(
                {"error": "No file uploaded. Please provide an Excel file."}, status=400
            )

        if not file.name.endswith((".xls", ".xlsx", ".csv")):
            return Response(
                {"error": "Unsupported file format. Please upload an Excel file."},
                status=400,
            )

        # Use openpyxl to read the Excel file
        try:
            # Step 3: Process the Excel file
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            emails_to_save = []
            invalid_rows = []
            duplicate_email = []

            # using a HashSet to query once (O(1)) instead of multiple queries (O(n)) in the loop.
            # changable in the future
            existing_emails = set(
                Email.objects.filter(campaign_id=campaign).values_list(
                    "email_address", flat=True
                )
            )

            for row_number, row in enumerate(
                sheet.iter_rows(min_row=2), start=2
            ):  # Start from the second row
                name = row[0].value
                recipient_email = row[1].value

                # Row validation
                if not recipient_email or "@" not in recipient_email:
                    invalid_rows.append(
                        {
                            "row": row_number,
                            "error": "Invalid or missing email address.",
                        }
                    )
                    continue

                # Skip if email already exists
                if recipient_email in existing_emails:
                    duplicate_email.append(recipient_email)
                    continue  # skips the email if its already there :)

                # Prepare the data for bulk create
                emails_to_save.append(
                    Email(
                        campaign_id=campaign,
                        name=name,
                        email_address=recipient_email,
                    )
                )

            # Step 4: Save emails to the database
            with transaction.atomic():
                Email.objects.bulk_create(emails_to_save)

            if invalid_rows:
                return Response(
                    {
                        "message": f"Emails saved successfully, but some rows were invalid.",
                        "invalid_rows": invalid_rows,
                    },
                    status=201,
                )

            if len(duplicate_email) != 0:
                return Response(
                    {
                        "message": f"Emails saved successfully, but some emails already exists in database under campaign ID {campaign_id}",
                        "duplicate email count": len(duplicate_email),
                        "duplicate emails": duplicate_email,
                    },
                    status=201,
                )

            return Response({"message": "Emails saved successfully!"}, status=201)

        except InvalidFileException:
            return Response(
                {"error": "Invalid Excel file. Please upload a valid Excel file."},
                status=400,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=500)


class SendEmailsView(APIView):
    @extend_schema(
        parameters=[
            OpenApiParameter(
                "campaign_id",
                type=int,
                location="query",
                required=True,
                description="ID of the campaign to send emails for.",
            ),
            OpenApiParameter(
                "email_template",
                type=int,
                location="query",
                required=True,
                description="Enter 1 for AutoSAD template or 2 for XCV AI template or 3 for AutoSAD.V2 or 4 for AutoSAD.V3 template",
                enum=["AutoSAD v1", "AutoSAD v2", "AutoSAD v3", "XCV AI"],
            ),
        ],
        request={
            "application/json": {
                "type": "object",
                "properties": {
                    "message": {
                        "type": "string",
                        "description": "Custom message to send in the email (optional). If not provided, the default message will be used.",
                    }
                },
                "example": {
                    "message": "Thank you for applying to the AUTOSAD Get Certified program. We're thrilled to have you on board and look forward to helping you gain the knowledge and credentials to excel in the AUTOSAD ecosystem. To finalize your enrollment and start your certification journey, simply click the link below to complete your registration process.",
                },
            }
        },
        responses={
            200: OpenApiResponse(
                description="Emails processed successfully.",
                examples={
                    "application/json": {
                        "message": "Emails processed.",
                        "details": {"sent": 10, "failed": 2},
                        "custom_message_used": "Thank you for applying to the AUTOSAD Get Certified program. We're thrilled to have you on board and look forward to helping you gain the knowledge and credentials to excel in the AUTOSAD ecosystem. To finalize your enrollment and start your certification journey, simply click the link below to complete your registration process.",
                    }
                },
            ),
            400: OpenApiResponse(
                description="Invalid or missing campaign ID.",
                examples={
                    "application/json": {"error": "Invalid or missing campaign ID."}
                },
            ),
            404: OpenApiResponse(
                description="No emails found for the given campaign.",
                examples={
                    "application/json": {
                        "error": "No emails found for the given campaign."
                    }
                },
            ),
            500: OpenApiResponse(
                description="Internal server error.",
                examples={"application/json": {"error": "Internal server error."}},
            ),
        },
        description="Send emails to all recipients associated with a specific campaign. Optionally provide a custom message.\n (Make sure the email list is uploaded for specific campaign ID before sending emails)",
    )
    def post(self, request):
        # Get campaign ID from query parameters
        campaign_id = request.query_params.get("campaign_id")
        email_template = request.query_params.get("email_template")

        if not campaign_id:
            return Response({"error": "Campaign ID is required."}, status=400)

        if not email_template:
            return Response({"error": "Email Template is required."}, status=400)

        try:
            campaign = Campaign.objects.get(campaign_id=campaign_id)
        except Campaign.DoesNotExist:
            return Response({"error": "Campaign not found."}, status=404)

        emails = Email.objects.filter(campaign_id=campaign)

        if not emails.exists():
            return Response(
                {"error": "No emails found for the given campaign."}, status=404
            )

        # Get the custom message from the request, or use the default
        custom_message = request.data.get("message")

        if not custom_message:
            custom_message = "Thank you for applying to the AUTOSAD Get Certified program. We're thrilled to have you on board and look forward to helping you gain the knowledge and credentials to excel in the AUTOSAD ecosystem. To finalize your enrollment and start your certification journey, simply click the link below to complete your registration process."
        success_count = 0
        failure_count = 0

        # Check if it's a built-in template or custom template
        builtin_templates = {
            "1": {"file": "autosad-temp-email.html", "subject": "Welcome to AUTOSAD Get Certified"},
            "2": {"file": "XCV_AI.html", "subject": "Welcome onboard to XCV AI"},
            "3": {"file": "autosad-temp-email2.html", "subject": "Welcome to AUTOSAD Get Certified"},
            "4": {"file": "autosad-email-temp-3.html", "subject": "Welcome to AUTOSAD"}
        }

        is_custom_template = False
        template_file = None
        template_subject = None
        custom_html_content = None

        if email_template in builtin_templates:
            # Built-in template
            template_info = builtin_templates[email_template]
            template_file = template_info["file"]
            template_subject = template_info["subject"]
        else:
            # Check if it's a custom template
            try:
                custom_template = EmailTemplate.objects.get(template_id=email_template)
                is_custom_template = True
                template_subject = custom_template.subject
                custom_html_content = custom_template.html_content
            except EmailTemplate.DoesNotExist:
                return Response({"error": "Template not found."}, status=400)

        for email in emails:
            try:
                if is_custom_template:
                    # For custom templates, check if it's a filename or direct content
                    if custom_html_content.endswith('.html'):
                        # It's a filename, use render_to_string like built-in templates
                        context = {
                            "name": email.name,
                            "message": custom_message,
                        }
                        html_content = render_to_string(custom_html_content, context)
                    else:
                        # Legacy: stored content directly in database
                        html_content = custom_html_content
                        if email.name:
                            html_content = html_content.replace('{{name}}', email.name)
                            html_content = html_content.replace('{{message}}', custom_message)
                else:
                    # For built-in templates, use render_to_string
                    context = {
                        "name": email.name,
                        "message": custom_message,
                    }
                    html_content = render_to_string(template_file, context)

                send_mail(
                    subject=template_subject,
                    message="",
                    from_email="info@autosad.ai",
                    recipient_list=[email.email_address],
                    html_message=html_content,
                )
                success_count += 1

            except Exception as e:
                failure_count += 1
                print(f"Failed to send email to {email.email_address}: {str(e)}")

        # for later purpose.

        # context = {'name': email.name, 'message': custom_message}
        # html_content = render_to_string(from_email, context)
        # email_msg = email.EmailMessage(
        #     subject,
        #     html_content,
        #     'info@autosad.ai',
        #     [email.email_address],
        #     connection=connection,  # Reuse connection here
        # )
        # email_msg.content_subtype = 'html'
        # email_msg.send()

        return Response(
            {
                "message": "Emails sent successfully!",
                "details": {
                    "sent": success_count,
                    "failed": failure_count,
                },
                "custom_message_used": custom_message,
            },
            status=200 if failure_count == 0 else 207,
        )  # 207: Multi-Status for partial success


class ListEmailView(APIView):
    @extend_schema(
        parameters=[
            OpenApiParameter(
                "campaign_id",
                type=int,
                location="query",
                required=True,
                description="ID of the campaign to list emails for.",
            )
        ],
        responses={
            200: OpenApiResponse(
                description="List of emails associated with the given campaign.",
                examples={
                    "application/json": [
                        {
                            "email_address": "john.doe@example.com",
                            "name": "John Doe",
                            "campaign_id": 1,
                            "added_at": "2025-01-01T12:00:00Z",
                        },
                        {
                            "email_address": "jane.smith@example.com",
                            "name": "Jane Smith",
                            "campaign_id": 2,
                            "added_at": "2025-01-10T14:00:00Z",
                        },
                    ]
                },
            ),
            400: OpenApiResponse(
                description="Invalid or missing campaign ID.",
                examples={
                    "application/json": {"error": "Invalid or missing campaign ID."}
                },
            ),
            404: OpenApiResponse(
                description="No emails found for the given campaign.",
                examples={
                    "application/json": {
                        "error": "No emails found for the given campaign."
                    }
                },
            ),
            500: OpenApiResponse(
                description="Internal server error.",
                examples={"application/json": {"error": "Internal server error."}},
            ),
        },
        description="List all emails stored in the database for a specific campaign.",
    )
    def get(self, request):
        # Get campaign ID from query parameters
        campaign_id = request.query_params.get("campaign_id")

        if not campaign_id:
            return Response({"error": "Campaign ID is required."}, status=400)

        try:
            campaign = Campaign.objects.get(campaign_id=campaign_id)
        except Campaign.DoesNotExist:
            return Response({"error": "Campaign not found."}, status=404)

        try:
            emails = Email.objects.filter(campaign_id=campaign)

            if not emails.exists():
                return Response(
                    {"error": "No emails found for the given campaign."}, status=404
                )

            serializer = EmailSerializer(emails, many=True)
            return Response(serializer.data, status=200)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class UpdateCampaignView(APIView):
    @extend_schema(
        parameters=[
            OpenApiParameter(
                "campaign_id",
                type=int,
                location="query",
                required=True,
                description="ID of the campaign to update.",
            ),
            OpenApiParameter(
                "campaign_name",
                type=str,
                location="query",
                required=True,
                description="New Campaign Name",
            ),
        ],
        responses={
            200: OpenApiResponse(
                description="Campaign name updated successfully.",
                examples={
                    "application/json": {
                        "campaign_id": 1,
                        "campaign_name": "Updated Campaign Name",
                        "created_at": "2025-01-27T10:00:00Z",
                        "updated_at": "2025-01-27T12:00:00Z",
                    }
                },
            ),
            400: OpenApiResponse(
                description="Invalid request or missing campaign ID.",
                examples={
                    "application/json": {"error": "Invalid or missing campaign ID."}
                },
            ),
            404: OpenApiResponse(
                description="Campaign not found.",
                examples={"application/json": {"error": "Campaign not found."}},
            ),
            500: OpenApiResponse(
                description="Internal server error.",
                examples={"application/json": {"error": "Internal server error."}},
            ),
        },
        description="Update the name of a specific campaign in the database.",
    )
    def post(self, request):
        campaign_id = request.query_params.get("campaign_id")
        updated_campaign_name = request.query_params.get("campaign_name")

        if not campaign_id or not updated_campaign_name:
            return Response(
                {"error": "Campaign ID and new campaign name are required."}, status=400
            )

        try:
            campaign = Campaign.objects.get(campaign_id=campaign_id)
        except Campaign.DoesNotExist:
            return Response({"error": "Campaign not found."}, status=404)

        try:
            campaign.campaign_name = updated_campaign_name
            campaign.save()

            return Response(
                {
                    "campaign_id": campaign.campaign_id,
                    "campaign_name": campaign.campaign_name,
                    "created_at": campaign.created_at,
                    "updated_at": campaign.updated_at,
                },
                status=200,
            )
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class DeleteCampaignView(APIView):
    @extend_schema(
        parameters=[
            OpenApiParameter(
                "campaign_id",
                type=int,
                location="query",
                required=False,
                description="ID of the campaign to delete.",
            )
        ],
        responses={
            200: OpenApiResponse(
                description="Campaign deleted successfully.",
                examples={
                    "application/json": {
                        "message": "Campaign and associated emails deleted successfully.",
                        "campaign_id": 1,
                    }
                },
            ),
            400: OpenApiResponse(
                description="Invalid request or missing campaign ID.",
                examples={
                    "application/json": {"error": "Invalid or missing campaign ID."}
                },
            ),
            404: OpenApiResponse(
                description="Campaign not found.",
                examples={"application/json": {"error": "Campaign not found."}},
            ),
            500: OpenApiResponse(
                description="Internal server error.",
                examples={"application/json": {"error": "Internal server error."}},
            ),
        },
        description="Delete a specific campaign and all its associated emails from the database.",
    )
    def post(self, request):
        campaign_id = request.query_params.get("campaign_id")

        if not campaign_id:
            return Response({"error": "Campaign ID is required."}, status=400)

        try:
            campaign = Campaign.objects.get(campaign_id=campaign_id)
        except Campaign.DoesNotExist:
            return Response({"error": "Campaign not found."}, status=404)

        try:
            campaign = Campaign.objects.get(campaign_id=campaign_id)
        except Campaign.DoesNotExist:
            return Response({"error": "Campaign not found."}, status=404)

        try:
            campaign.delete()

            return Response(
                {
                    "message": "Campaign and associated emails deleted successfully.",
                    "campaign_id": campaign_id,
                },
                status=200,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=500)


class DeleteEmailView(APIView):
    @extend_schema(
        parameters=[
            OpenApiParameter(
                "email_add",
                type=str,
                location="query",
                required=False,
                description="Email address from the database to delete.",
            ),
            OpenApiParameter(
                "campaign_id",
                type=int,
                location="query",
                required=True,
                description="Campaign ID of the campagain to delete.",
            ),
        ],
        responses={
            200: OpenApiResponse(
                description="Email address deleted successfully.",
                examples={
                    "application/json": {
                        "Message": "Campaign and associated emails deleted successfully.",
                        "Email Address": "sample@mail.com",
                        "Campaign ID": "1",
                    }
                },
            ),
            400: OpenApiResponse(
                description="Invalid request or missing email address.",
            ),
            404: OpenApiResponse(
                description="Email address not found.",
                examples={
                    "application/json": {"error": "Email Database empty or not found."}
                },
            ),
            500: OpenApiResponse(
                description="Internal server error.",
                examples={"application/json": {"error": "Internal server error."}},
            ),
        },
        description="Delete an email address and all its associated ID, campaign_id and other data.",
    )
    def post(self, request):
        email_add = request.query_params.get("email_add")
        campaign_id = request.query_params.get("campaign_id")

        if not email_add:
            return Response({"error": "Email address is required."}, status=400)

        if not campaign_id:
            return Response({"error": "Campaign ID is required."}, status=400)

        # try:
        #     email = Email.objects.get(email_address=email_add)
        # except Email.DoesNotExist:
        #     return Response({"error": "Email Database empty or not found."}, status=404)

        # try:
        #     campaign_id = Email.objects.get(campaign_id=campaign_id)
        # except Email.DoesNotExist:
        #     return Response(
        #         {"error": "Campaign ID not found."},
        #         status=404
        #     )

        try:
            email = Email.objects.get(email_address=email_add, campaign_id=campaign_id)
        except Email.DoesNotExist:
            return Response(
                {
                    "error": f"Email: [{email_add}] with campaign ID: {campaign_id} not found."
                },
                status=404,
            )

        try:
            email.delete()
            return Response(
                {
                    "Message": "Campaign and associated emails deleted successfully.",
                    "Email Address": email_add,
                    "Campaign ID": campaign_id,
                },
                status=200,
            )
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class UpdateEmailView(APIView):
    @extend_schema(
        parameters=[
            OpenApiParameter(
                "campaign_id",
                type=int,
                location="query",
                required=True,
                description="ID of the campaign to update.",
            ),
            OpenApiParameter(
                "email_id",
                type=str,
                location="query",
                required=True,
                description="ID of email to update.",
            ),
            OpenApiParameter(
                "email_add",
                type=str,
                location="query",
                required=True,
                description="New email address to update",
            ),
        ],
        responses={
            200: OpenApiResponse(
                description="Email address updated successfully.",
                examples={
                    "application/json": {
                        "campaign_id": 1,
                        "email_id": "6",
                        "New email address": "new_mail@mail.com",
                        "updated_at": "2025-01-27T12:00:00Z",
                    }
                },
            ),
            400: OpenApiResponse(
                description="Invalid request or missing parameters.",
                examples={
                    "application/json": {
                        "error": "Campaign ID and email ID are required."
                    }
                },
            ),
            404: OpenApiResponse(
                description="Email not found.",
                examples={"application/json": {"error": "Email not found."}},
            ),
            500: OpenApiResponse(
                description="Internal server error.",
                examples={"application/json": {"error": "Internal server error."}},
            ),
        },
        description="Update the email address for a specific campaign in the database.",
    )
    def post(self, request):
        campaign_id = request.query_params.get("campaign_id")
        email_id = request.query_params.get("email_id")
        email_add = request.query_params.get("email_add")

        if not campaign_id or not email_id:
            return Response(
                {"error": "Campaign ID and email address are required."}, status=400
            )

        if not email_add:
            return Response({"error": "Email address is required."}, status=400)

        try:
            email = Email.objects.get(campaign_id=campaign_id, email_id=email_id)
        except Email.DoesNotExist:
            return Response(
                {
                    "error": f"Email with ID [{email_id}] not found in Campaign [{campaign_id}]."
                },
                status=404,
            )

        try:
            email.email_add = email_add
            email.save()  # saves the email from the ORM

            return Response(
                {
                    "campaign_id": 1,
                    "email_id": "6",
                    "New email address": "new_mail@mail.com",
                    "updated_at": "2025-01-27T12:00:00Z",
                },
                status=200,
            )
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class EmailTemplatePreviewView(APIView):
    @extend_schema(
        parameters=[
            OpenApiParameter(
                "template_id",
                type=str,
                location="query",
                required=True,
                description="Template ID to preview (1-4 for built-in, or custom template ID)",
            )
        ],
        responses={
            200: OpenApiResponse(
                description="Email template HTML content",
                examples={
                    "application/json": {
                        "template_name": "AutoSAD v1",
                        "subject": "Welcome to AUTOSAD Get Certified",
                        "html_content": "<html>...</html>"
                    }
                },
            ),
            400: OpenApiResponse(
                description="Invalid template ID",
                examples={"application/json": {"error": "Invalid template ID"}},
            ),
            500: OpenApiResponse(
                description="Internal server error",
                examples={"application/json": {"error": "Template not found"}},
            ),
        },
        description="Preview email template content by template ID",
    )
    def get(self, request):
        template_id = request.query_params.get("template_id")

        if not template_id:
            return Response({"error": "Template ID is required"}, status=400)

        print(f"Template preview requested for ID: {template_id}")

        try:
            # Check if it's a built-in template (1-4)
            builtin_templates = {
                "1": {
                    "name": "AutoSAD v1",
                    "subject": "Welcome to AUTOSAD Get Certified",
                    "file": "autosad-temp-email.html"
                },
                "2": {
                    "name": "XCV AI", 
                    "subject": "Welcome onboard to XCV AI",
                    "file": "XCV_AI.html"
                },
                "3": {
                    "name": "AutoSAD v2",
                    "subject": "Welcome to AUTOSAD Get Certified", 
                    "file": "autosad-temp-email2.html"
                },
                "4": {
                    "name": "AutoSAD v3",
                    "subject": "Welcome to AUTOSAD",
                    "file": "autosad-email-temp-3.html"
                }
            }

            if template_id in builtin_templates:
                template_info = builtin_templates[template_id]
                # Render template with sample data for preview
                context = {
                    "name": "John Doe",
                    "message": "Thank you for applying to the AUTOSAD Get Certified program. We're thrilled to have you on board and look forward to helping you gain the knowledge and credentials to excel in the AUTOSAD ecosystem. To finalize your enrollment and start your certification journey, simply click the link below to complete your registration process."
                }
                html_content = render_to_string(template_info["file"], context)
                
                return Response({
                    "template_name": template_info["name"],
                    "subject": template_info["subject"], 
                    "html_content": html_content
                }, status=200)
            else:
                # Check if it's a custom template
                try:
                    # Convert template_id to integer for custom templates
                    template_id_int = int(template_id)
                    print(f"Looking for custom template with ID: {template_id_int}")
                    custom_template = EmailTemplate.objects.get(template_id=template_id_int)
                    print(f"Found custom template: {custom_template.template_name}")</old_str>
                    
                    # Load template content from file if it's a filename, otherwise use stored content
                    if custom_template.html_content.endswith('.html'):
                        # It's a filename, load from file
                        context = {
                            "name": "John Doe",
                            "message": "Thank you for applying to the AUTOSAD Get Certified program. We're thrilled to have you on board and look forward to helping you gain the knowledge and credentials to excel in the AUTOSAD ecosystem. To finalize your enrollment and start your certification journey, simply click the link below to complete your registration process."
                        }
                        html_content = render_to_string(custom_template.html_content, context)
                    else:
                        # Legacy: stored content directly in database
                        html_content = custom_template.html_content
                        html_content = html_content.replace('{{name}}', 'John Doe')
                        html_content = html_content.replace('{{message}}', 'Thank you for applying to the AUTOSAD Get Certified program. We\'re thrilled to have you on board and look forward to helping you gain the knowledge and credentials to excel in the AUTOSAD ecosystem. To finalize your enrollment and start your certification journey, simply click the link below to complete your registration process.')
                    
                    return Response({
                        "template_name": custom_template.template_name,
                        "subject": custom_template.subject,
                        "html_content": html_content
                    }, status=200)
                except (EmailTemplate.DoesNotExist, ValueError):
                    return Response({"error": "Template not found"}, status=404)

        except Exception as e:
            return Response({"error": f"Template not found: {str(e)}"}, status=500)


class EmailTemplateListView(APIView):
    @extend_schema(
        responses={
            200: OpenApiResponse(
                description="List of all email templates (built-in and custom)",
                examples={
                    "application/json": [
                        {
                            "template_id": "1",
                            "template_name": "AutoSAD v1",
                            "type": "builtin"
                        },
                        {
                            "template_id": "5",
                            "template_name": "My Custom Template",
                            "type": "custom"
                        }
                    ]
                },
            ),
        },
        description="Get all available email templates including built-in and custom ones",
    )
    def get(self, request):
        try:
            templates = []
            
            # Add built-in templates
            builtin_templates = [
                {"template_id": "1", "template_name": "AutoSAD v1", "type": "builtin"},
                {"template_id": "2", "template_name": "XCV AI", "type": "builtin"},
                {"template_id": "3", "template_name": "AutoSAD v2", "type": "builtin"},
                {"template_id": "4", "template_name": "AutoSAD v3", "type": "builtin"}
            ]
            templates.extend(builtin_templates)
            
            # Add custom templates with proper error handling
            try:
                # Try to fetch custom templates directly - Django will handle table existence
                custom_templates = EmailTemplate.objects.all().order_by('-created_at')
                for template in custom_templates:
                    templates.append({
                        "template_id": str(template.template_id),
                        "template_name": template.template_name,
                        "type": "custom"
                    })
                        
            except Exception as db_error:
                print(f"Error fetching custom templates: {db_error}")
                # Continue with just built-in templates if database error occurs
            
            return Response(templates, status=200)
            
        except Exception as e:
            print(f"Error in EmailTemplateListView: {e}")
            # Return built-in templates even if there's an error
            builtin_templates = [
                {"template_id": "1", "template_name": "AutoSAD v1", "type": "builtin"},
                {"template_id": "2", "template_name": "XCV AI", "type": "builtin"},
                {"template_id": "3", "template_name": "AutoSAD v2", "type": "builtin"},
                {"template_id": "4", "template_name": "AutoSAD v3", "type": "builtin"}
            ]
            return Response(builtin_templates, status=200)


class EmailTemplateCreateView(APIView):
    @extend_schema(
        request=EmailTemplateSerializer,
        responses={
            201: EmailTemplateSerializer,
            400: OpenApiResponse(description="Bad Request"),
        },
        description="Create a new custom email template",
    )
    def post(self, request):
        try:
            serializer = EmailTemplateSerializer(data=request.data)
            if serializer.is_valid():
                template = serializer.save()
                
                # Save the HTML content as a file in the templates folder
                import os
                template_filename = f"custom_template_{template.template_id}.html"
                template_path = os.path.join(settings.BASE_DIR, 'mailer', 'templates', template_filename)
                
                with open(template_path, 'w', encoding='utf-8') as f:
                    f.write(template.html_content)
                
                # Update the template record to store the filename
                template.html_content = template_filename
                template.save()
                
                return Response(serializer.data, status=201)
            return Response(serializer.errors, status=400)
        except Exception as e:
            return Response({"error": str(e)}, status=500)